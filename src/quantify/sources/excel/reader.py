"""Excel file reader supporting both .xls and .xlsx formats."""

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


@dataclass
class ColumnRange:
    """Parsed column range specification."""

    column: str  # Column letter (e.g., "K")
    start_row: int  # Starting row number (e.g., 2)
    end_row: int | None  # Ending row (None means "to last data row")


class ExcelReader:
    """Reads values from Excel files.

    Supports:
    - .xlsx files via openpyxl
    - .xls files via xlrd
    """

    def __init__(self, file_path: Path) -> None:
        """Initialize reader.

        Args:
            file_path: Path to the Excel file.
        """
        self._file_path = file_path
        self._workbook: Any = None
        self._is_xlsx = file_path.suffix.lower() == ".xlsx"

    def _ensure_workbook(self) -> None:
        """Lazy-load the workbook."""
        if self._workbook is not None:
            return

        if self._is_xlsx:
            import openpyxl

            self._workbook = openpyxl.load_workbook(
                self._file_path,
                data_only=True,  # Get computed values, not formulas
                read_only=True,
            )
        else:
            import xlrd

            self._workbook = xlrd.open_workbook(str(self._file_path))

    @staticmethod
    def parse_column_range(range_str: str) -> ColumnRange:
        """Parse a column range string like 'K2:K' or 'K2:K100'.

        Args:
            range_str: Range string (e.g., "K2:K", "AB5:AB100")

        Returns:
            Parsed ColumnRange.

        Raises:
            ValueError: If range format is invalid.
        """
        # Pattern: COLUMN + START_ROW + ":" + COLUMN (+ optional END_ROW)
        pattern = r"^([A-Z]+)(\d+):([A-Z]+)(\d*)$"
        match = re.match(pattern, range_str.upper())
        if not match:
            raise ValueError(f"Invalid column range format: {range_str}")

        start_col, start_row, end_col, end_row = match.groups()

        if start_col != end_col:
            raise ValueError(f"Column range must be same column: {range_str}")

        return ColumnRange(
            column=start_col,
            start_row=int(start_row),
            end_row=int(end_row) if end_row else None,
        )

    @staticmethod
    def column_letter_to_index(column: str) -> int:
        """Convert column letter(s) to 0-based index.

        Args:
            column: Column letter(s) like "A", "Z", "AA", "AB"

        Returns:
            0-based column index.
        """
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1  # Convert to 0-based

    def get_tab_sum(self, tab_name: str, column_range: ColumnRange) -> float:
        """Get sum of values in a column range for a specific tab.

        Args:
            tab_name: Name of the worksheet/tab.
            column_range: Parsed column range.

        Returns:
            Sum of numeric values in the range.
        """
        self._ensure_workbook()

        if self._is_xlsx:
            return self._get_sum_xlsx(tab_name, column_range)
        else:
            return self._get_sum_xls(tab_name, column_range)

    def _get_sum_xlsx(self, tab_name: str, col_range: ColumnRange) -> float:
        """Get sum from xlsx file using openpyxl."""
        try:
            sheet = self._workbook[tab_name]
        except KeyError:
            return 0.0

        col_idx = self.column_letter_to_index(col_range.column) + 1  # openpyxl is 1-based
        total = 0.0

        # Determine end row
        end_row = col_range.end_row or sheet.max_row

        for row in range(col_range.start_row, end_row + 1):
            cell = sheet.cell(row=row, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                total += float(cell.value)

        return total

    def _get_sum_xls(self, tab_name: str, col_range: ColumnRange) -> float:
        """Get sum from xls file using xlrd."""
        import xlrd

        try:
            sheet = self._workbook.sheet_by_name(tab_name)
        except xlrd.biffh.XLRDError:
            return 0.0

        col_idx = self.column_letter_to_index(col_range.column)  # xlrd is 0-based
        total = 0.0

        # Determine end row (xlrd uses 0-based indexing)
        start_row = col_range.start_row - 1  # Convert to 0-based
        end_row = (col_range.end_row - 1) if col_range.end_row else sheet.nrows - 1

        for row in range(start_row, min(end_row + 1, sheet.nrows)):
            try:
                cell = sheet.cell(row, col_idx)
                # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error
                if cell.ctype == 2:  # NUMBER type - don't use date conversion
                    total += float(cell.value)
            except IndexError:
                continue

        return total

    def get_monthly_sums(
        self,
        tab_name: str,
        date_range: ColumnRange,
        value_range: ColumnRange,
    ) -> dict[int, float]:
        """Get sums grouped by month from date+value columns.

        Parses dates in DD.MM.YYYY format (German locale).

        Args:
            tab_name: Name of the worksheet/tab.
            date_range: Parsed column range for dates.
            value_range: Parsed column range for values.

        Returns:
            Dict mapping month number (1-12) to sum of values.
        """
        self._ensure_workbook()

        if self._is_xlsx:
            return self._get_monthly_sums_xlsx(tab_name, date_range, value_range)
        else:
            return self._get_monthly_sums_xls(tab_name, date_range, value_range)

    def _parse_date_string(self, date_str: str) -> datetime | None:
        """Parse a date string in DD.MM.YYYY format.

        Args:
            date_str: Date string like "12.01.2026" or "5.3.2024".

        Returns:
            datetime object or None if parsing fails.
        """
        try:
            return datetime.strptime(date_str.strip(), "%d.%m.%Y")
        except (ValueError, AttributeError):
            return None

    def _get_monthly_sums_xlsx(
        self,
        tab_name: str,
        date_range: ColumnRange,
        value_range: ColumnRange,
    ) -> dict[int, float]:
        """Get monthly sums from xlsx file using openpyxl."""
        try:
            sheet = self._workbook[tab_name]
        except KeyError:
            return {}

        date_col_idx = self.column_letter_to_index(date_range.column) + 1  # openpyxl is 1-based
        value_col_idx = self.column_letter_to_index(value_range.column) + 1

        monthly_totals: dict[int, float] = {}

        # Use the same start row for both, end at the larger end_row or max_row
        start_row = max(date_range.start_row, value_range.start_row)
        end_row = min(
            date_range.end_row or sheet.max_row,
            value_range.end_row or sheet.max_row,
        )

        for row in range(start_row, end_row + 1):
            date_cell = sheet.cell(row=row, column=date_col_idx)
            value_cell = sheet.cell(row=row, column=value_col_idx)

            # Get date - can be datetime or string
            date_value = date_cell.value
            if date_value is None:
                continue

            month: int | None = None
            if isinstance(date_value, datetime):
                month = date_value.month
            elif isinstance(date_value, str):
                parsed = self._parse_date_string(date_value)
                if parsed:
                    month = parsed.month

            if month is None:
                continue

            # Get value
            if value_cell.value is not None and isinstance(value_cell.value, (int, float)):
                monthly_totals[month] = monthly_totals.get(month, 0.0) + float(value_cell.value)

        return monthly_totals

    def _get_monthly_sums_xls(
        self,
        tab_name: str,
        date_range: ColumnRange,
        value_range: ColumnRange,
    ) -> dict[int, float]:
        """Get monthly sums from xls file using xlrd."""
        import xlrd

        try:
            sheet = self._workbook.sheet_by_name(tab_name)
        except xlrd.biffh.XLRDError:
            return {}

        date_col_idx = self.column_letter_to_index(date_range.column)  # xlrd is 0-based
        value_col_idx = self.column_letter_to_index(value_range.column)

        monthly_totals: dict[int, float] = {}

        # Use the same start row for both (convert to 0-based)
        start_row = max(date_range.start_row, value_range.start_row) - 1
        end_row = min(
            (date_range.end_row - 1) if date_range.end_row else sheet.nrows - 1,
            (value_range.end_row - 1) if value_range.end_row else sheet.nrows - 1,
        )

        for row in range(start_row, min(end_row + 1, sheet.nrows)):
            try:
                date_cell = sheet.cell(row, date_col_idx)
                value_cell = sheet.cell(row, value_col_idx)
            except IndexError:
                continue

            # Get month from date cell
            month: int | None = None

            # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error
            if date_cell.ctype == 3:  # DATE type
                try:
                    dt = xlrd.xldate_as_datetime(date_cell.value, self._workbook.datemode)
                    month = dt.month
                except (ValueError, TypeError):
                    pass
            elif date_cell.ctype == 1:  # TEXT type
                parsed = self._parse_date_string(str(date_cell.value))
                if parsed:
                    month = parsed.month

            if month is None:
                continue

            # Get value
            if value_cell.ctype == 2:  # NUMBER type
                monthly_totals[month] = monthly_totals.get(month, 0.0) + float(value_cell.value)

        return monthly_totals

    def get_available_tabs(self) -> list[str]:
        """Get list of available tab names.

        Returns:
            List of worksheet names.
        """
        self._ensure_workbook()

        if self._is_xlsx:
            return list(self._workbook.sheetnames)
        else:
            return list(self._workbook.sheet_names())

    def close(self) -> None:
        """Close the workbook and release resources."""
        if self._workbook is not None:
            if self._is_xlsx:
                self._workbook.close()
            self._workbook = None
